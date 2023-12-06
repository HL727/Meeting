from django.http import HttpResponse
from django.utils.text import slugify
from django.utils.timezone import now
from drf_extra_fields.fields import Base64FileField
from rest_framework import serializers
from io import BytesIO


class HybridFileField(Base64FileField):

    def to_internal_value(self, data):
        from django.core.exceptions import ValidationError
        try:
            return super().to_internal_value( data)
        except ValidationError:
            return serializers.FileField.to_internal_value(self, data)


class HybridExcelFileField(HybridFileField):

    ALLOWED_TYPES = ('xls', 'xlsx')


class ExcelFileInputSerializer(serializers.Serializer):

    file = HybridExcelFileField()
    ignore_empty = serializers.BooleanField(required=False, default=True)

    def validate(self, attrs):

        ignore_empty = attrs.get('ignore_empty', True)
        cur_file = attrs['file']

        if cur_file.name.lower().endswith('.xlsx'):
            result = self._validate_xlxs(cur_file, ignore_empty=ignore_empty)
        elif cur_file.name.lower().endswith('.xls'):
            result = self._validate_xls(cur_file, ignore_empty=ignore_empty)
        else:
            return []

        return self._remove_empty_lines(result)

    def _validate_xlxs(self, cur_file, ignore_empty):
        result = []
        from openpyxl import load_workbook
        sheet = load_workbook(cur_file, read_only=True).worksheets[0]

        for row in sheet.rows:
            cur = ['' if c.value is None else c.value for c in row]
            if not ignore_empty or any(cur):
                result.append(cur)
        return result

    def _validate_xls(self, cur_file, ignore_empty):
        result = []
        from xlrd.book import open_workbook_xls
        sheet = open_workbook_xls(file_contents=cur_file.read()).sheet_by_index(0)

        for row in range(sheet.nrows):
            cur = [sheet.cell(row, col).value for col in range(sheet.ncols)]
            cur = ['' if c is None else c for c in cur]
            if not ignore_empty or any(cur):
                result.append(cur)
        return result

    def _remove_empty_lines(self, result):

        i = len(result) - 1
        while i > 0:
            if any(result[i]):
                return result[:i + 1]
            i -= 1

        return result


class ExcelCreateFileSerializer(serializers.Serializer):

    headers = serializers.ListField(child=serializers.CharField(), required=False, allow_empty=True)
    rows = serializers.ListField(child=serializers.ListField())
    filename = serializers.CharField(required=False)

    def validate(self, attrs):

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active

        if attrs.get('headers'):
            ws.append(attrs['headers'])

        for row in attrs['rows']:
            ws.append(row)

        for column_cells in ws.columns:
            length = max(len(str(cell.value or '0')) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(60, length)

        fd = BytesIO()
        wb.save(fd)
        fd.seek(0)

        filename = attrs.get('filename', '').strip()
        if not filename:
            filename = 'download-{}'.format(now().replace(microsecond=0)).replace(':', '')

        return {
            'filename': '{}.xlsx'.format(slugify(filename)),
            'fd': fd,
        }

    def get_file_extension(self, filename, decoded_file):
        ext = filename.lower().rsplit('.')[-1]
        if ext:
            return ext
        return 'bin'

    def get_response(self):

        self.is_valid(raise_exception=True)

        data = self.validated_data

        response = HttpResponse(data['fd'])
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response['Content-Disposition'] = 'attachment; filename="{}"'.format(data['filename'])
        return response


def rest_filterset_factory(model, fields='__all__'):
    """
    like `django_filters.filterset.filterset_factory` but with rest framework specific fields, e.g. for
    datetimes
    """
    from django_filters.rest_framework import FilterSet
    meta = type(str('Meta'), (object,), {'model': model, 'fields': fields})
    filterset = type(str('%sFilterSet' % model._meta.object_name),
                     (FilterSet,), {'Meta': meta})
    return filterset
